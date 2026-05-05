"""Excel .xlsx converter: workbook → canonical markdown with per-sheet preview.

Uses openpyxl in read-only mode for memory-efficient streaming over large
workbooks. Each sheet renders as its own ## section with a preview table
(50 rows × 20 cols by default). The original .xlsx is preserved as a
sidecar at raw/xlsx/<id>.xlsx; the markdown body is a preview only.

Canonical ID per WIKI § 6.1: xlsx-<author>-<year>-<short> when metadata
allows, else xlsx-<sha256-prefix-12> hash fallback (mirrors PDF / DOCX).
"""

from __future__ import annotations

import hashlib
import re
import shutil
from datetime import datetime, timezone
from pathlib import Path
from zipfile import BadZipFile

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from gateway import frontmatter as fm
from gateway import paths, validator
from gateway.converters.base import ConversionError, Converter

_PREVIEW_ROWS = 50
_PREVIEW_COLS = 20

_NON_ALNUM = re.compile(r"[^a-z0-9]+")


def _now_iso() -> str:
    return datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")


def _slugify_short(value: str, max_words: int = 4) -> str:
    cleaned = _NON_ALNUM.sub("-", value.lower()).strip("-")
    parts = [p for p in cleaned.split("-") if p]
    return "-".join(parts[:max_words])


def _hash_id_from_bytes(data: bytes, max_chars: int = 12) -> str:
    digest = hashlib.sha256(data).hexdigest()[:max_chars]
    return f"xlsx-{digest}"


def _build_id(author: str, title: str, year: str, raw_bytes: bytes) -> str:
    if author and year and title:
        author_slug = _slugify_short(author.split(",")[0], max_words=2)
        title_slug = _slugify_short(title, max_words=3)
        if author_slug and title_slug:
            return f"xlsx-{author_slug}-{year}-{title_slug}"
    return _hash_id_from_bytes(raw_bytes)


def _md_escape_cell(value) -> str:
    if value is None:
        return ""
    return (
        str(value)
        .replace("\\", "\\\\")
        .replace("|", "\\|")
        .replace("\r", " ")
        .replace("\n", " ")
    )


def _render_table(headers: list, rows: list[list]) -> str:
    if not headers:
        return ""
    width = len(headers)
    out = ["| " + " | ".join(_md_escape_cell(h) for h in headers) + " |",
           "| " + " | ".join(["---"] * width) + " |"]
    for row in rows:
        padded = list(row[:width])
        while len(padded) < width:
            padded.append("")
        out.append("| " + " | ".join(_md_escape_cell(c) for c in padded) + " |")
    return "\n".join(out)


def _render_sheet_section(sheet) -> tuple[str, int, int]:
    """Render one worksheet as a markdown section.

    Returns (markdown, total_rows, total_cols). total_rows excludes the
    header row.
    """
    rows_iter = sheet.iter_rows(values_only=True)
    try:
        headers = list(next(rows_iter))
    except StopIteration:
        return f"## {sheet.title}\n\n_(empty sheet)_", 0, 0

    # Strip trailing all-None columns from headers (openpyxl pads to max col)
    while headers and headers[-1] is None:
        headers.pop()
    if not headers:
        return f"## {sheet.title}\n\n_(empty sheet)_", 0, 0

    total_cols = len(headers)
    truncated_cols = total_cols > _PREVIEW_COLS
    display_headers = headers[:_PREVIEW_COLS]

    display_rows: list[list] = []
    total_rows = 0
    for row in rows_iter:
        # Skip rows that are entirely None (openpyxl yields these for tables
        # that span past the data region).
        if all(c is None for c in row):
            continue
        total_rows += 1
        if len(display_rows) < _PREVIEW_ROWS:
            row_list = list(row[:_PREVIEW_COLS]) if truncated_cols else list(row[:total_cols])
            display_rows.append(row_list)

    parts = [f"## {sheet.title}",
             "",
             f"{total_rows} data rows × {total_cols} columns.",
             "",
             _render_table(display_headers, display_rows)]
    notes: list[str] = []
    if total_rows > _PREVIEW_ROWS:
        notes.append(f"{total_rows - _PREVIEW_ROWS} additional rows omitted from preview")
    if truncated_cols:
        notes.append(f"{total_cols - _PREVIEW_COLS} additional columns omitted from preview")
    if notes:
        parts.append("")
        parts.append(f"_{'; '.join(notes)}. Full data in sidecar._")

    return "\n".join(parts), total_rows, total_cols


class XlsxConverter(Converter):
    type_name = "xlsx"

    def detect(self, source: str) -> bool:
        if source.startswith(("http://", "https://")):
            return False
        try:
            return Path(source).suffix.lower() == ".xlsx"
        except (TypeError, ValueError):
            return False

    def convert(self, source: str) -> str:
        path = Path(source).expanduser().resolve()
        if not path.exists():
            raise ConversionError(f"xlsx not found: {path}")

        raw_bytes = path.read_bytes()
        try:
            wb = load_workbook(filename=str(path), read_only=True, data_only=True)
        except (InvalidFileException, BadZipFile, KeyError, ValueError) as e:
            raise ConversionError(f"openpyxl failed on {path}: {e}") from e

        sheet_sections: list[str] = []
        sheet_count = 0
        total_data_rows = 0
        sheet_summaries: list[dict] = []
        for ws in wb.worksheets:
            section, n_rows, n_cols = _render_sheet_section(ws)
            sheet_sections.append(section)
            if n_cols > 0:
                sheet_count += 1
                total_data_rows += n_rows
                sheet_summaries.append({
                    "name": ws.title,
                    "rows": n_rows,
                    "columns": n_cols,
                })

        if sheet_count == 0:
            wb.close()
            raise ConversionError(f"no extractable content in {path}")

        cp = wb.properties
        author = (cp.creator or "").strip()
        title = (cp.title or path.stem).strip()
        created = cp.created
        year = created.strftime("%Y") if isinstance(created, datetime) else ""
        wb.close()

        body_parts = [f"# {title}",
                      "",
                      f"Workbook with **{sheet_count}** sheet(s) and {total_data_rows} total data rows.",
                      "",
                      *sheet_sections]
        body = "\n\n".join(p for p in body_parts if p != "")
        body = body.rstrip("\n") + "\n"

        canonical_id = _build_id(author, cp.title or "", year, raw_bytes)

        sidecar_dir = paths.raw_dir_for("xlsx")
        sidecar_dir.mkdir(parents=True, exist_ok=True)
        sidecar_target = sidecar_dir / f"{canonical_id}.xlsx"
        if not sidecar_target.exists():
            shutil.copy2(path, sidecar_target)

        front: dict = {
            "id": canonical_id,
            "type": "xlsx",
            "title": title,
            "url": "",
            "authors": _split_authors(author),
            "ingested_at": _now_iso(),
            "content_hash": validator.compute_content_hash(body),
            "source_path": str(sidecar_target.relative_to(paths.knowledge_root())),
            "domains": [],
            "nlm_corpus_ids": [],
            "wiki_pages": [],
            "meta": {
                "sheet_count": sheet_count,
                "sheets": sheet_summaries,
                "total_data_rows": total_data_rows,
                "extraction_tool": "openpyxl",
                "original_filename": path.name,
            },
        }
        if year:
            front["published_at"] = year
        return fm.serialize(front, body)


def _split_authors(value: str) -> list[str]:
    if not value:
        return []
    parts = [p.strip() for p in re.split(r"[,;]| and ", value)]
    return [p for p in parts if p]
