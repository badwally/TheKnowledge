"""Tests for the Excel .xlsx converter (M31)."""

from __future__ import annotations

from datetime import datetime
from pathlib import Path

import pytest
from openpyxl import Workbook

from gateway import converters
from gateway import frontmatter as fm
from gateway import paths
from gateway.converters import xlsx as xlsx_mod
from gateway.converters.base import ConversionError


@pytest.fixture(autouse=True)
def _reset_registry():
    converters.reset_registry_for_tests()
    yield
    converters.reset_registry_for_tests()


def _build_xlsx(
    path: Path,
    *,
    sheets: list[tuple[str, list[list]]],
    creator: str = "",
    title: str = "",
    created: datetime | None = None,
) -> None:
    """Write a .xlsx file. sheets is [(sheet_name, [[row1], [row2], ...])]."""
    wb = Workbook()
    # Remove the default sheet
    default = wb.active
    wb.remove(default)
    for name, rows in sheets:
        ws = wb.create_sheet(name)
        for row in rows:
            ws.append(row)
    if creator:
        wb.properties.creator = creator
    if title:
        wb.properties.title = title
    if created is not None:
        wb.properties.created = created
    wb.save(str(path))


# --- detect ----------------------------------------------------------------


def test_xlsx_detect_accepts_xlsx_paths():
    c = xlsx_mod.XlsxConverter()
    assert c.detect("/tmp/foo.xlsx")
    assert c.detect("foo.XLSX")


def test_xlsx_detect_rejects_other_extensions_and_urls():
    c = xlsx_mod.XlsxConverter()
    assert not c.detect("/tmp/foo.xls")  # legacy not supported
    assert not c.detect("/tmp/foo.csv")
    assert not c.detect("https://example.com/foo.xlsx")


# --- convert ---------------------------------------------------------------


def test_xlsx_convert_single_sheet_writes_sidecar(kb_root, tmp_path):
    src = tmp_path / "data.xlsx"
    _build_xlsx(
        src,
        sheets=[
            ("Sales", [
                ["region", "qty", "amount"],
                ["NA", 12, 150.00],
                ["EU", 7, 98.50],
            ]),
        ],
        creator="Andrew Grant",
        title="Q1 Budget",
        created=datetime(2026, 4, 28),
    )

    text = xlsx_mod.XlsxConverter().convert(str(src))
    front, body = fm.parse(text)

    assert front["type"] == "xlsx"
    assert front["title"] == "Q1 Budget"
    assert front["authors"] == ["Andrew Grant"]
    assert front["published_at"] == "2026"
    assert front["meta"]["sheet_count"] == 1
    assert front["meta"]["total_data_rows"] == 2
    assert front["meta"]["sheets"][0]["name"] == "Sales"
    assert front["meta"]["sheets"][0]["columns"] == 3

    assert "## Sales" in body
    assert "| region | qty | amount |" in body
    assert "| NA | 12 |" in body

    sidecar = paths.raw_dir_for("xlsx") / f"{front['id']}.xlsx"
    assert sidecar.exists()
    assert sidecar.read_bytes() == src.read_bytes()


def test_xlsx_convert_multi_sheet_renders_each_section(kb_root, tmp_path):
    src = tmp_path / "multi.xlsx"
    _build_xlsx(
        src,
        sheets=[
            ("Sheet1", [["a", "b"], [1, 2]]),
            ("Notes", [["title", "comment"], ["item 1", "ok"]]),
        ],
    )
    text = xlsx_mod.XlsxConverter().convert(str(src))
    front, body = fm.parse(text)

    assert front["meta"]["sheet_count"] == 2
    assert "## Sheet1" in body
    assert "## Notes" in body


def test_xlsx_convert_uses_author_year_title_id(kb_root, tmp_path):
    src = tmp_path / "data.xlsx"
    _build_xlsx(
        src,
        sheets=[("S", [["h"], ["v"]])],
        creator="Smith",
        title="The Big Idea",
        created=datetime(2025, 1, 1),
    )
    text = xlsx_mod.XlsxConverter().convert(str(src))
    front, _ = fm.parse(text)
    assert front["id"].startswith("xlsx-smith-2025-")


def test_xlsx_convert_falls_back_to_hash_id_without_metadata(kb_root, tmp_path):
    src = tmp_path / "data.xlsx"
    _build_xlsx(src, sheets=[("S", [["h"], ["v"]])])
    text = xlsx_mod.XlsxConverter().convert(str(src))
    front, _ = fm.parse(text)
    suffix = front["id"][len("xlsx-"):]
    assert len(suffix) == 12
    assert all(c in "0123456789abcdef" for c in suffix)


def test_xlsx_convert_truncates_large_rows_with_note(kb_root, tmp_path):
    src = tmp_path / "big.xlsx"
    rows = [["col1", "col2"]]
    rows.extend([[f"r{i}", f"v{i}"] for i in range(75)])
    _build_xlsx(src, sheets=[("Big", rows)])

    text = xlsx_mod.XlsxConverter().convert(str(src))
    front, body = fm.parse(text)

    assert front["meta"]["total_data_rows"] == 75
    assert "25 additional rows omitted from preview" in body


def test_xlsx_convert_truncates_wide_columns_with_note(kb_root, tmp_path):
    src = tmp_path / "wide.xlsx"
    headers = [f"c{i}" for i in range(25)]
    data = [f"v{i}" for i in range(25)]
    _build_xlsx(src, sheets=[("Wide", [headers, data])])

    text = xlsx_mod.XlsxConverter().convert(str(src))
    front, body = fm.parse(text)

    assert front["meta"]["sheets"][0]["columns"] == 25
    assert "5 additional columns omitted from preview" in body


def test_xlsx_convert_escapes_pipe_in_cell(kb_root, tmp_path):
    src = tmp_path / "pipes.xlsx"
    _build_xlsx(src, sheets=[("S", [["name", "note"], ["foo", "a|b|c"]])])
    text = xlsx_mod.XlsxConverter().convert(str(src))
    _, body = fm.parse(text)
    assert "a\\|b\\|c" in body


def test_xlsx_convert_idempotent_id_for_same_content(kb_root, tmp_path):
    """Two byte-identical .xlsx files produce the same canonical ID."""
    src1 = tmp_path / "first.xlsx"
    src2 = tmp_path / "second.xlsx"
    _build_xlsx(src1, sheets=[("S", [["h"], ["v"]])])
    src2.write_bytes(src1.read_bytes())

    f1, _ = fm.parse(xlsx_mod.XlsxConverter().convert(str(src1)))
    f2, _ = fm.parse(xlsx_mod.XlsxConverter().convert(str(src2)))
    assert f1["id"] == f2["id"]


def test_xlsx_convert_missing_file_raises(kb_root):
    with pytest.raises(ConversionError):
        xlsx_mod.XlsxConverter().convert("/does/not/exist.xlsx")


def test_xlsx_convert_empty_workbook_raises(kb_root, tmp_path):
    src = tmp_path / "empty.xlsx"
    wb = Workbook()
    # Default sheet has no headers/rows
    default = wb.active
    default.title = "Sheet1"
    wb.save(str(src))
    with pytest.raises(ConversionError):
        xlsx_mod.XlsxConverter().convert(str(src))


def test_xlsx_convert_corrupt_file_raises(kb_root, tmp_path):
    src = tmp_path / "corrupt.xlsx"
    src.write_bytes(b"not a real xlsx file at all")
    with pytest.raises(ConversionError):
        xlsx_mod.XlsxConverter().convert(str(src))


# --- registry --------------------------------------------------------------


def test_dispatch_xlsx_path_routes_to_xlsx_converter():
    c = converters.dispatch("/tmp/data.xlsx")
    assert isinstance(c, xlsx_mod.XlsxConverter)


def test_dispatch_xls_legacy_format_has_no_converter():
    """Legacy .xls is intentionally unsupported in v1."""
    from gateway.converters import NoConverterError
    with pytest.raises(NoConverterError):
        converters.dispatch("/tmp/old.xls")
